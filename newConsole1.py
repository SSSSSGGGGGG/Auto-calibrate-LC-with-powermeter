# -*- coding: utf-8 -*-
"""
Created on Wed Feb 14 11:28:11 2024

@author: Laboratorio
"""

import initialPhotodiode as iniP
import closePhotodiode as clP
import initialKLC
import closeKLC
import numpy as np
import measurePower2 as mp
from KLCCommandLib import *
import numpy as np
from openpyxl import Workbook
import time
import asyncio
import nest_asyncio
nest_asyncio.apply()

# =============================================================================
#               initial all the devices with proper parameters
# =============================================================================
# initial photodiode
# to link the photodiode (pd), the parameter a=1 (not change it) is to call initialPhtodiode.py
# newrange=0 of initialPhtodiode.py, means we choose "auto" scal to pd
a=1
newrange=0
iniP.iniPhotodiode (a,newrange)

# initial KLC 
# here we also use a=1 to call KLC initial function
# If we need to restart KLC, we should restart the kernal (in the python console)
# To check the handle value of KLC, which is necessary for the following KLC functions.
handle=initialKLC.initialKLC(a)
print("KLC´s handle:",handle)


# =============================================================================
# 
# =============================================================================
# vols2=[0,1,2,3,4,5]
# x=np.arange(0,0.8,0.1)
# y=np.arange(0.8,2.1,0.01)
# z=np.arange(2.1,5.1,0.1)
vols2=np.arange(0,5.1,0.1)
# The file will save all the measurements
config1=input("what is the configuration (please input c or p):")
filename=input("Please input a file name:")
if config1=="c":
    
    # create a worksheet, and save tiles for required each cols, I set col1 to 3 for crossed, and 5-7 for para
    workbook=Workbook()
    worksheet=workbook.active
    worksheet.cell(1, 1, 'Vol_crossed')
    worksheet.cell(1, 2, 'I (W)')
    # worksheet.cell(1, 3, 'T (ms)')
    worksheet.cell(1, 20, 'Vol_para')
    worksheet.cell(1, 21, 'I (W)')
    
    for k in range(len(vols2)):
        worksheet.cell(k+2, 1, vols2[k])
        worksheet.cell(k+2, 20, vols2[k])
        #worksheet.cell(k+3, 9, volsOutput[k])
    workbook.save(filename+r".xlsx")

# config value will be vavriable for judging the type of configuration.

f=1000  # !!!!! set the frequency to the enabled channel
mode=2 # 1 continuous; 2 cycle.
cyclenumber=1 #number of cycles 1~ 2147483648.s
delay=1000  # !!! the sample intervals[ms] 1~ 2147483648
precycle_rest=0  # the delay time before the cycle start[ms] 0~ 2147483648.
hdl=0  


async def sendVoltage(vols):
    l=len(vols)
    volarr =  (c_float * len(vols))(*vols)
    if(klcSetOutputLUT(hdl, volarr, l)<0):
        print("klcSetOutputLUT failed")
        
    if(klcSetOutputLUTParams(hdl, mode, cyclenumber, delay, precycle_rest)<0):
        print("klcSetOutputLUTParams failed")
    
    print("----------------Current output voltage is:",vols)
    
    task=asyncio.create_task(photodiode())
   
    if(klcStartLUTOutput(hdl)<0):
        print("klcStartLUTOutput failed")
    await task
    await asyncio.sleep(1)
    
def my_fun(s, i=[0]):
    print(s)
    i[0] += 1
    return i[0]   
async def photodiode():
    sleep=0.7  # !!!! it couldnot be very small, we should check it in practice, unit is "s"
    t=my_fun("number")
    print("------",t)
    mp.measureAction(sleep,config1,filename,t)   
    await asyncio.sleep(1.5)  
   
def main():        
    length=len(vols2) #!!!!
    count=1
    while(count<length+1):  
        vols1=[vols2[count-1]]        
        asyncio.run(sendVoltage(vols1))
        count+=1

if __name__=="__main__":
    main()

# =============================================================================
#                    close all the devices
# =============================================================================
# input y will close all device, any other input will keep pd work, but KLC will be closed anyway
end=input("Do you want to close your photodiode (please input y for yes): ")
clP.close(end)
closeKLC.closeKLC()