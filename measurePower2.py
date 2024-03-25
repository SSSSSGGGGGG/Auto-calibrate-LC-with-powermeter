# -*- coding: utf-8 -*-
"""
Created on Wed Jan 31 14:53:16 2024

@author: Laboratorio
"""


import win32com.client
import time
import traceback
from openpyxl import  load_workbook
# =============================================================================
#     
# =============================================================================

def measureAction(sleep1,config,filename,t):     
    try:
      OphirCOM = win32com.client.Dispatch("OphirLMMeasurement.CoLMMeasurement")      
      DeviceList = OphirCOM.ScanUSB()[0] # [0] is the index for device number
      DeviceHandle = OphirCOM.OpenUSBDevice(DeviceList)     
      OphirCOM.StartStream(DeviceHandle, 0)	
	  
      time.sleep(sleep1)				# exposure time, wait a little for data, the unit is second and it it is very small the output vlaue will overlap with each other
      data = OphirCOM.GetData(DeviceHandle, 0)    # start measuring
      
      #OphirCOM.StopAllStreams()
     
      cols=len(data[0])         
      dataT=[data[0],data[1]] 
      print("-------measurements for each voltage:",dataT)
      # Open an exist excel
      workbook=load_workbook(filename+r".xlsx")
      worksheet=workbook.active
      if config=="c":  # When config is string c, the data will be save at cols 1-3
         
        j=0
        while(j<cols):
         worksheet.cell(t+1, j+2, dataT[0][j])                
         j+=1
                      
        workbook.save(filename+r".xlsx")
      else:     # When config is not string c, the data will be save at cols 5-7
          j=0
          while(j<cols):
           worksheet.cell(t+1, j+21, dataT[0][j])                
           j+=1
                        
          workbook.save(filename+r".xlsx")
      if len(data[0]) > 0:		# if any data available, print the first one from the batch
        print('photodiode Reading = {0}W, TimeStamp = {1}ms, Status = {2} '.format(data[0][0] ,data[1][0] ,data[2][0]))
    
      else:
          print('\nNo Sensor attached to {0} !!!'.format(DeviceList))
      
      
    except OSError as err:
     print("OS error: {0}".format(err))
    except:
     traceback.print_exc()
     
    
